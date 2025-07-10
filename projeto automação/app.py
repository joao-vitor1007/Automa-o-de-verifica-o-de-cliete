import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import keyboard

# Abrir planilha de clientes
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

# Pegando o primeiro cliente (linha 2)
linha = next(pagina_clientes.iter_rows(min_row=2, max_row=2, values_only=True))
nome, valor, cpf, vencimento = linha

# Abrir/Cria planilha de fechamento
try:
    planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
    pagina_fechamento = planilha_fechamento['Sheet1']
except FileNotFoundError:
    planilha_fechamento = openpyxl.Workbook()
    pagina_fechamento = planilha_fechamento.active
    pagina_fechamento.title = "Sheet1"
    pagina_fechamento.append(['Nome', 'Valor', 'CPF', 'Vencimento', 'Status', 'Data Pagamento', 'Método'])

# Iniciar navegador
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')
sleep(3)

# Verificar cliente
campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
campo_pesquisa.clear()
sleep(1)
campo_pesquisa.send_keys(cpf)
sleep(1)

botao_pesquisar = driver.find_element(By.XPATH, "//button[@class = 'btn btn-custom btn-lg btn-block mt-3']")
botao_pesquisar.click()
sleep(3)

status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")

if status.text == 'em dia':
    data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']").text
    metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']").text
    data_pagamento_limpo = data_pagamento.text.split()[3]
    metodo_pagamento_limpo = metodo_pagamento.text.split()[3]
    pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])
   
else:
    pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente', '', ''])

# Salvar a planilha com os dados finais
planilha_fechamento.save('planilha fechamento.xlsx')

# Aguardar tecla para sair
print("Pressione a tecla 'Esc' para fechar o navegador...")
while True:  
    if keyboard.is_pressed('esc'): 
        print("Tecla 'Esc' pressionada. Fechando o navegador...")
        break 

driver.quit()

    